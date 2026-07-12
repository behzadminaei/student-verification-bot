#!/usr/bin/env python3
"""Import students from an Excel sheet into the SQLite students table.

Expected columns:
  Full Name, Student-Number, Password, url

Student-Number is stored as both student_number and username.
url is stored as exam_url.
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook

REQUIRED_HEADERS = ("Full Name", "Student-Number", "Password", "url")


def _header_map(header_row: tuple[object, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip()
        if name:
            mapping[name] = index
    missing = [name for name in REQUIRED_HEADERS if name not in mapping]
    if missing:
        raise SystemExit(
            "Missing required Excel columns: " + ", ".join(missing)
        )
    return mapping


def _cell_str(row: tuple[object, ...], index: int) -> str:
    value = row[index] if index < len(row) else None
    if value is None:
        return ""
    return str(value).strip()


def import_students(xlsx_path: Path, db_path: Path) -> None:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration as exc:
        raise SystemExit("Excel file is empty") from exc

    columns = _header_map(tuple(header))
    conn = sqlite3.connect(db_path)
    inserted = 0
    skipped = 0

    try:
        for row_number, row in enumerate(rows, start=2):
            if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            full_name = _cell_str(row, columns["Full Name"])
            student_number = _cell_str(row, columns["Student-Number"])
            password = _cell_str(row, columns["Password"])
            exam_url = _cell_str(row, columns["url"])

            if not full_name or not student_number or not password or not exam_url:
                print(
                    f"Row {row_number}: skipped (missing required field)",
                    file=sys.stderr,
                )
                skipped += 1
                continue

            try:
                conn.execute(
                    """
                    INSERT INTO students (
                        full_name,
                        student_number,
                        username,
                        password,
                        exam_url
                    )
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (
                        full_name,
                        student_number,
                        student_number,
                        password,
                        exam_url,
                    ),
                )
                inserted += 1
            except sqlite3.IntegrityError:
                print(
                    f"Row {row_number}: skipped (duplicate student_number "
                    f"{student_number})",
                    file=sys.stderr,
                )
                skipped += 1

        conn.commit()
    finally:
        conn.close()
        workbook.close()

    print(f"Inserted {inserted} student(s); skipped {skipped}.")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import students from Excel into SQLite."
    )
    parser.add_argument("xlsx_path", type=Path, help="Path to .xlsx file")
    parser.add_argument("db_path", type=Path, help="Path to students.db")
    args = parser.parse_args()

    if not args.xlsx_path.is_file():
        raise SystemExit(f"Excel file not found: {args.xlsx_path}")
    if not args.db_path.is_file():
        raise SystemExit(f"Database file not found: {args.db_path}")

    import_students(args.xlsx_path, args.db_path)


if __name__ == "__main__":
    main()
